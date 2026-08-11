"""动态生成调查数据 Excel 导入模板。

模板完全依据数据库 information_schema 元数据生成，因此 sheet 名、列名、
必填列与 `/api/survey/excel-import` 的校验规则保持一致。
模板按虫种生成：虫种来自 pest_registry 注册表，表归属按"表名包含虫种名"匹配。
示例行取各表最新一条真实记录；表为空时回退到按类型的占位示例。
"""

from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.db.pool import quote_identifier
from backend.db.postgres import ensure_pool
from backend.logging_config import get_logger
from backend.services.pest_registry import validate_pest_type
from backend.services.survey_excel_import import (
    fetch_survey_table_metadata,
    is_backend_generated_column,
)


logger = get_logger(__name__)

# 取示例行时的"最新"排序键优先级：调查表按调查日期，流水表按事件时间，兜底 id
EXAMPLE_ROW_ORDER_COLUMNS = ("调查日期", "事件时间", "id")

REQUIRED_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
REQUIRED_FONT = Font(color="842029", bold=True)
HEADER_FILL = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")
HEADER_FONT = Font(bold=True)


def build_example_value(column: Any) -> Any:
    """根据列数据类型生成示例值（表内无真实数据时的兜底）。"""

    data_type = (column.data_type or "").lower()
    udt_name = (column.udt_name or "").lower()

    if data_type == "date":
        return date(2026, 5, 1).isoformat()
    if data_type.startswith("timestamp"):
        return "2026-05-01 08:00:00"
    if data_type in {"integer", "bigint", "smallint"}:
        return 1
    if data_type == "boolean" or udt_name == "bool":
        return "是"
    if data_type == "numeric" or data_type == "decimal" or udt_name == "numeric":
        return 1.0
    return "示例值"


def build_example_values(table_meta: Any, example_row: dict[str, Any] | None) -> list[Any]:
    """生成示例行：优先使用真实数据行，NULL 与自动生成列留空，缺列回退到类型占位。"""

    ordered_columns = sorted(
        table_meta.columns.values(),
        key=lambda col: col.ordinal_position,
    )
    example_row = example_row or {}
    values = []
    for column in ordered_columns:
        if column.is_auto_generated or is_backend_generated_column(table_meta, column.name):
            values.append(None)
            continue
        if column.name in example_row:
            values.append(example_row[column.name])
        else:
            values.append(build_example_value(column))
    return values


def is_required_column(column: Any) -> bool:
    """与导入规则保持一致：非空、无默认值、非自动生成。"""

    if column.is_auto_generated:
        return False
    if column.is_nullable:
        return False
    if column.default.strip():
        return False
    return True


def build_import_template_bytes(
    metadata: dict[str, Any],
    example_rows: dict[str, dict[str, Any]] | None = None,
) -> bytes:
    """根据表元数据生成多 sheet 导入模板。

    example_rows 可选：表名 → 真实数据行，用作示例行；缺省回退到类型占位示例。
    """

    workbook = Workbook(write_only=False)
    # 移除默认 sheet，按实际表动态创建
    default_sheet = workbook.active
    if default_sheet is not None:
        workbook.remove(default_sheet)

    for table_meta in metadata.values():
        worksheet = workbook.create_sheet(title=table_meta.name)
        ordered_columns = sorted(
            table_meta.columns.values(),
            key=lambda col: col.ordinal_position,
        )

        # 表头行
        headers = [column.name for column in ordered_columns]
        worksheet.append(headers)
        for col_index, column in enumerate(ordered_columns, start=1):
            cell = worksheet.cell(row=1, column=col_index)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            if is_required_column(column):
                cell.fill = REQUIRED_FILL
                cell.font = REQUIRED_FONT
                cell.comment = Comment("必填列，导入时不能为空", "系统")

        # 示例行
        example_values = build_example_values(
            table_meta,
            (example_rows or {}).get(table_meta.name),
        )
        worksheet.append(example_values)
        for col_index, column in enumerate(ordered_columns, start=1):
            cell = worksheet.cell(row=2, column=col_index)
            if is_required_column(column):
                cell.fill = REQUIRED_FILL
                cell.font = REQUIRED_FONT

        # 自动调整列宽（粗略估算）
        for col_index, column in enumerate(ordered_columns, start=1):
            header_len = len(column.name)
            example_len = len(str(example_values[col_index - 1]))
            width = min(max(header_len, example_len) + 4, 40)
            worksheet.column_dimensions[get_column_letter(col_index)].width = width

        logger.debug("生成模板 sheet: %s, 列数: %d", table_meta.name, len(ordered_columns))

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def filter_template_metadata(
    metadata: dict[str, Any],
    pest_type: str,
) -> dict[str, Any]:
    """按虫种过滤模板元数据，只保留表名包含该虫种名的可导入表。"""

    validated_pest = validate_pest_type(pest_type)
    filtered = {
        table_name: table_meta
        for table_name, table_meta in metadata.items()
        if validated_pest in table_meta.name
    }
    if not filtered:
        raise ValueError(f"{validated_pest} 没有可导入的数据表")
    return filtered


async def fetch_latest_example_rows(
    connection: Any,
    metadata: dict[str, Any],
) -> dict[str, dict[str, Any]]:
    """每张表取最新一行真实数据作为模板示例；空表不出现在返回值中。"""

    example_rows: dict[str, dict[str, Any]] = {}
    for table_meta in metadata.values():
        order_column = next(
            (name for name in EXAMPLE_ROW_ORDER_COLUMNS if name in table_meta.columns),
            None,
        )
        qualified_table = (
            f"{quote_identifier(table_meta.schema_name)}.{quote_identifier(table_meta.name)}"
        )
        order_clause = (
            f"ORDER BY {quote_identifier(order_column)} DESC" if order_column else ""
        )
        row = await connection.fetchrow(
            f"SELECT * FROM {qualified_table} {order_clause} LIMIT 1"
        )
        if row is not None:
            example_rows[table_meta.name] = dict(row)
    return example_rows


async def generate_import_template_bytes(pest_type: str) -> bytes:
    """异步获取数据库元数据并生成指定虫种的模板字节流。"""

    pool = await ensure_pool()
    async with pool.acquire() as connection:
        metadata = await fetch_survey_table_metadata(connection)
        filtered = filter_template_metadata(metadata, pest_type)
        example_rows = await fetch_latest_example_rows(connection, filtered)

    logger.info("生成导入模板: 虫种=%s 共 %d 张表", pest_type, len(filtered))
    return build_import_template_bytes(filtered, example_rows)
