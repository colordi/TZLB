from __future__ import annotations

from datetime import datetime
from typing import Any

from openpyxl import Workbook

from backend.db.postgres import ensure_pool
from backend.services.data_export.metadata import (
    _build_filter_params,
    fetch_export_table_metadata,
    fetch_pest_export_metadata,
    fetch_pest_export_metadata_filtered,
    get_export_table_meta,
    list_export_tables,
    list_pest_export_types,
)
from backend.services.data_export.types import DataExportArtifact, XLSX_MEDIA_TYPE
from backend.services.data_export.workbook import (
    append_summary_sheet,
    append_table_sheet,
    build_export_filename,
    build_unique_sheet_names,
    validate_export_schema,
    workbook_to_bytes,
)


async def export_all_tables() -> DataExportArtifact:
    exported_at = datetime.now()
    workbook = Workbook(write_only=True)

    pool = await ensure_pool()
    async with pool.acquire() as connection:
        tables = await fetch_export_table_metadata(connection)
        if not tables:
            raise ValueError("没有可导出的 survey 或 ledger 表或视图")

        sheet_names = build_unique_sheet_names(tables)
        append_summary_sheet(workbook, tables, sheet_names, exported_at)
        for table in tables:
            await append_table_sheet(
                workbook,
                connection,
                table,
                sheet_names[(table.schema_name, table.table_name)],
            )

    return DataExportArtifact(
        filename=build_export_filename("调查数据导出", exported_at),
        media_type=XLSX_MEDIA_TYPE,
        content=workbook_to_bytes(workbook),
    )


async def export_single_table(schema_name: str, table_name: str) -> DataExportArtifact:
    exported_at = datetime.now()
    workbook = Workbook(write_only=True)

    pool = await ensure_pool()
    async with pool.acquire() as connection:
        table = await get_export_table_meta(connection, schema_name, table_name)
        sheet_name = build_unique_sheet_names([table])[(table.schema_name, table.table_name)]
        await append_table_sheet(workbook, connection, table, sheet_name)

    return DataExportArtifact(
        filename=build_export_filename(f"{schema_name}_{table_name}", exported_at),
        media_type=XLSX_MEDIA_TYPE,
        content=workbook_to_bytes(workbook),
    )


async def export_pest_type(
    pest_type: str,
    year: str | None = None,
    generation: str | None = None,
) -> DataExportArtifact:
    exported_at = datetime.now()
    workbook = Workbook(write_only=True)

    pool = await ensure_pool()
    async with pool.acquire() as connection:
        pest_metas = await fetch_pest_export_metadata(connection, pest_type=pest_type)
        if not pest_metas or not pest_metas[0].tables:
            raise ValueError(f"虫种不存在或无数据：{pest_type}")

        tables = list(pest_metas[0].tables)
        sheet_names = build_unique_sheet_names(tables)

        filter_parts = [f"虫种：{pest_type}"]
        if year:
            filter_parts.append(f"年份={year}")
        if generation:
            filter_parts.append(f"世代={generation}")

        append_summary_sheet(
            workbook,
            tables,
            sheet_names,
            exported_at,
            range_description="，".join(filter_parts),
        )

        for table in tables:
            conditions, params = _build_filter_params(table, year, generation)
            where_clause = " AND ".join(conditions) if conditions else None
            await append_table_sheet(
                workbook,
                connection,
                table,
                sheet_names[(table.schema_name, table.table_name)],
                where_clause=where_clause,
                where_params=tuple(params),
            )

    filter_suffix = ""
    if year or generation:
        parts = []
        if year:
            parts.append(f"年份{year}")
        if generation:
            parts.append(f"世代{generation}")
        filter_suffix = f"_{'-'.join(parts)}"

    return DataExportArtifact(
        filename=build_export_filename(f"{pest_type}{filter_suffix}", exported_at),
        media_type=XLSX_MEDIA_TYPE,
        content=workbook_to_bytes(workbook),
    )
