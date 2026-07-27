from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from backend.services.survey_import.metadata import is_backend_generated_column
from backend.services.survey_import.parsers import (
    is_blank,
    normalize_header_value,
    parse_cell_value,
)
from backend.services.survey_import.types import (
    USE_DEFAULT,
    PreparedRow,
    PreparedSheet,
    TableMeta,
)


def sheet_has_data(sheet: Any) -> bool:
    for row in sheet.iter_rows(values_only=True):
        if any(not is_blank(value) for value in row):
            return True
    return False


def get_header(sheet: Any) -> list[str]:
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return [normalize_header_value(value) for value in first_row]


def build_import_plan(content: bytes, metadata: dict[str, TableMeta]) -> list[PreparedSheet]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"无法读取 Excel 文件：{exc}") from exc

    sheets: list[PreparedSheet] = []
    for worksheet in workbook.worksheets:
        if not sheet_has_data(worksheet):
            continue

        sheet_result = PreparedSheet(
            sheet_name=worksheet.title,
            schema_name=None,
            table_name=worksheet.title,
        )
        table_meta = metadata.get(worksheet.title)
        if table_meta is None:
            sheet_result.errors.append(
                f"sheet 名称必须与 survey 或 ledger 中的可写表完全一致：{worksheet.title}"
            )
            sheets.append(sheet_result)
            continue

        sheet_result.schema_name = table_meta.schema_name
        if not table_meta.conflict_columns:
            sheet_result.errors.append("目标表缺少主键或唯一键，不能安全跳过重复记录")
            sheets.append(sheet_result)
            continue

        headers = get_header(worksheet)
        if validate_headers(headers, table_meta, sheet_result):
            sheets.append(sheet_result)
            continue

        header_indexes = {
            header: index
            for index, header in enumerate(headers)
            if header and header in table_meta.columns
        }
        ignored_auto_columns = {
            header
            for header in header_indexes
            if table_meta.columns[header].is_auto_generated
            or is_backend_generated_column(table_meta, header)
        }

        validate_required_columns(header_indexes, table_meta, sheet_result)
        if sheet_result.errors:
            sheets.append(sheet_result)
            continue

        parse_rows(worksheet, table_meta, header_indexes, ignored_auto_columns, sheet_result)
        sheets.append(sheet_result)

    return sheets


def validate_headers(
    headers: list[str],
    table_meta: TableMeta,
    sheet_result: PreparedSheet,
) -> bool:
    if not any(headers):
        sheet_result.errors.append("第 1 行必须是数据库列名")
        return True

    seen_headers: set[str] = set()
    for header in headers:
        if not header:
            continue
        if header in seen_headers:
            sheet_result.errors.append(f"列名重复：{header}")
        seen_headers.add(header)
        if header not in table_meta.columns:
            sheet_result.errors.append(f"目标表不存在列：{header}")

    return bool(sheet_result.errors)


def validate_required_columns(
    header_indexes: dict[str, int],
    table_meta: TableMeta,
    sheet_result: PreparedSheet,
) -> None:
    missing_columns = []
    for column in table_meta.columns.values():
        if (
            column.is_auto_generated
            or is_backend_generated_column(table_meta, column.name)
            or column.is_nullable
            or column.has_default
        ):
            continue
        if column.name not in header_indexes:
            missing_columns.append(column.name)

    if missing_columns:
        sheet_result.errors.append(f"缺少必填列：{', '.join(missing_columns)}")


def parse_rows(
    worksheet: Any,
    table_meta: TableMeta,
    header_indexes: dict[str, int],
    ignored_auto_columns: set[str],
    sheet_result: PreparedSheet,
) -> None:
    ordered_headers = [
        header
        for header, _ in sorted(header_indexes.items(), key=lambda item: item[1])
        if header not in ignored_auto_columns
    ]

    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(not is_blank(value) for value in row):
            continue

        sheet_result.row_count += 1
        row_values: dict[str, Any] = {}
        row_errors: list[str] = []
        for header in ordered_headers:
            column = table_meta.columns[header]
            index = header_indexes[header]
            raw_value = row[index] if index < len(row) else None
            try:
                parsed_value = parse_cell_value(column, raw_value)
            except ValueError as exc:
                row_errors.append(f"{header}：{exc}")
                continue

            if parsed_value is None and not column.is_nullable and not column.has_default:
                row_errors.append(f"{header} 不能为空")
                continue
            if parsed_value is USE_DEFAULT:
                continue
            row_values[header] = parsed_value

        missing_conflict_columns = [
            column
            for column in table_meta.conflict_columns
            if is_blank(row_values.get(column))
        ]
        if missing_conflict_columns:
            row_errors.append(f"冲突键字段不能为空：{', '.join(missing_conflict_columns)}")

        if row_errors:
            sheet_result.errors.append(f"第 {row_number} 行：{'；'.join(row_errors)}")
            continue

        conflict_values = tuple(row_values[column] for column in table_meta.conflict_columns)
        sheet_result.rows.append(
            PreparedRow(
                row_number=row_number,
                values=row_values,
                conflict_values=conflict_values,
            )
        )
        sheet_result.valid_rows += 1


def mark_file_duplicates(sheet_result: PreparedSheet) -> None:
    seen_keys: set[tuple[Any, ...]] = set()
    for row in sheet_result.rows:
        if row.conflict_values in seen_keys:
            row.skipped_duplicate = True
            sheet_result.skipped_duplicate_rows += 1
            sheet_result.warnings.append(
                f"第 {row.row_number} 行与本文件前面记录重复，已跳过"
            )
            continue
        seen_keys.add(row.conflict_values)
