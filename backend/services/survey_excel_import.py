from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any

import asyncpg
from openpyxl import load_workbook

from backend.db.postgres import ensure_pool, quote_identifier
from backend.logging_config import get_logger


logger = get_logger(__name__)


ALLOWED_SCHEMAS = ("survey", "ledger")
AUTO_DEFAULT_MARKERS = ("nextval(", "generated")
EXCEL_EPOCH = date(1899, 12, 30)
EXCEL_DATETIME_EPOCH = datetime.combine(EXCEL_EPOCH, datetime.min.time())
RECHECK_ABNORMAL_EVENT_TYPE = "复查异常"

# 没有业务唯一键的流水表，冲突键统一硬编码为 (编号, 事件类型, 事件时间)，
# 插入时走 INSERT ... WHERE NOT EXISTS（不支持 ON CONFLICT）。
LEDGER_CONFLICT_COLUMNS = {
    ("ledger", "美国白蛾问题点位事件流水表"): ("编号", "事件类型", "事件时间"),
    ("ledger", "国槐尺蠖问题点位事件流水表"): ("编号", "事件类型", "事件时间"),
    ("ledger", "春尺蠖问题点位事件流水表"): ("编号", "事件类型", "事件时间"),
}
# id 既非 identity 也无默认值的流水表，由后端按 MAX(id)+1 分配。
BACKEND_GENERATED_ID_TABLES = {
    ("ledger", "美国白蛾问题点位事件流水表"),
    ("ledger", "国槐尺蠖问题点位事件流水表"),
}
# 历史对比纠正规则：(schema, 表名) -> (历史分组键, 下派类事件类型集合)。
# 用户录入下派类事件但同组已存在历史事件时，纠正为"复查异常"。
LEDGER_HISTORY_RULES = {
    ("ledger", "美国白蛾问题点位事件流水表"): (("编号", "年份", "世代"), {"调查下派"}),
    ("ledger", "国槐尺蠖问题点位事件流水表"): (
        ("编号", "年份", "世代"),
        {"历史预警下派", "幼虫调查下派"},
    ),
    ("ledger", "春尺蠖问题点位事件流水表"): (
        ("编号", "年份"),
        {"历史预警下派", "成虫调查下派", "幼虫调查下派"},
    ),
    ("ledger", "其他害虫问题点位事件流水表"): (("编号", "虫害类型", "年份"), {"调查下派"}),
}
LOCALITY_FIELD = "属地"
EVENT_TYPE_FIELD = "事件类型"
DAMAGED_PLANT_COUNT_FIELD = "受害株数"
DAMAGE_LEVEL_FIELD = "危害程度"
EMPTY_LOCALITY_LABEL = "未填写"
UNDAMAGED_LEVELS = frozenset({"", "白", "无需防治"})


@dataclass(frozen=True)
class ColumnMeta:
    name: str
    data_type: str
    udt_name: str
    is_nullable: bool
    default: str
    ordinal_position: int
    is_identity: bool = False
    enum_labels: tuple[str, ...] = ()

    @property
    def is_auto_generated(self) -> bool:
        normalized_default = self.default.lower()
        return self.is_identity or any(
            marker in normalized_default for marker in AUTO_DEFAULT_MARKERS
        )

    @property
    def has_default(self) -> bool:
        return self.default.strip() != ""


@dataclass(frozen=True)
class TableMeta:
    schema_name: str
    name: str
    columns: dict[str, ColumnMeta]
    conflict_columns: tuple[str, ...]
    supports_on_conflict: bool = True


@dataclass
class PreparedRow:
    row_number: int
    values: dict[str, Any]
    conflict_values: tuple[Any, ...]
    skipped_duplicate: bool = False


@dataclass
class PreparedSheet:
    sheet_name: str
    schema_name: str | None
    table_name: str | None
    row_count: int = 0
    valid_rows: int = 0
    inserted_rows: int = 0
    skipped_duplicate_rows: int = 0
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    rows: list[PreparedRow] = field(default_factory=list)


class UseColumnDefault:
    pass


USE_DEFAULT = UseColumnDefault()


def normalize_header_value(value: Any) -> str:
    return str(value or "").strip()


def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def parse_excel_date(value: Any) -> date | None:
    if is_blank(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, bool):
        raise ValueError("日期不能是布尔值")
    if isinstance(value, int) or isinstance(value, float):
        if isinstance(value, float) and not value.is_integer():
            raise ValueError("日期序列号必须是整数")
        return EXCEL_EPOCH + timedelta(days=int(value))

    text = str(value).strip()
    try:
        return date.fromisoformat(text[:10])
    except ValueError as exc:
        raise ValueError("日期格式必须是 YYYY-MM-DD 或 Excel 日期") from exc


def parse_excel_datetime(value: Any) -> datetime | None:
    if is_blank(value):
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, bool):
        raise ValueError("时间不能是布尔值")
    if isinstance(value, int) or isinstance(value, float):
        return EXCEL_DATETIME_EPOCH + timedelta(seconds=round(float(value) * 86400))

    text = str(value).strip()
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        try:
            return datetime.combine(date.fromisoformat(text[:10]), datetime.min.time())
        except ValueError as exc:
            raise ValueError("时间格式必须是 YYYY-MM-DD 或 ISO 日期时间") from exc


def parse_integer(value: Any) -> int | None:
    if is_blank(value):
        return None
    if isinstance(value, bool):
        raise ValueError("整数不能是布尔值")
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if not value.is_integer():
            raise ValueError("整数列不能包含小数")
        return int(value)

    text = str(value).strip()
    if not text.removeprefix("-").isdigit():
        raise ValueError("整数列只能填写整数")
    return int(text)


def parse_text(value: Any) -> str | None:
    if is_blank(value):
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_cell_value(column: ColumnMeta, value: Any) -> Any:
    if is_blank(value):
        if not column.is_nullable and column.has_default:
            return USE_DEFAULT
        return None

    if column.data_type == "date":
        return parse_excel_date(value)
    if column.data_type.startswith("timestamp"):
        return parse_excel_datetime(value)
    if column.data_type in {"integer", "bigint", "smallint"}:
        return parse_integer(value)
    text = parse_text(value)
    if text is not None and column.enum_labels and text not in column.enum_labels:
        raise ValueError(f"取值必须是：{'、'.join(column.enum_labels)}")
    return text


async def fetch_survey_table_metadata(connection: asyncpg.Connection) -> dict[str, TableMeta]:
    column_rows = await connection.fetch(
        """
        SELECT
            c.table_schema,
            c.table_name,
            c.ordinal_position,
            c.column_name,
            c.data_type,
            c.udt_name,
            c.is_nullable,
            c.is_identity,
            COALESCE(column_default, '') AS column_default
        FROM information_schema.columns AS c
        JOIN information_schema.tables AS t
          ON t.table_schema = c.table_schema
         AND t.table_name = c.table_name
        WHERE c.table_schema = ANY($1::text[])
          AND t.table_type = 'BASE TABLE'
        ORDER BY c.table_schema, c.table_name, c.ordinal_position
        """,
        list(ALLOWED_SCHEMAS),
    )
    constraint_rows = await connection.fetch(
        """
        SELECT
            tc.table_schema,
            tc.table_name,
            tc.constraint_name,
            tc.constraint_type,
            ARRAY_AGG(kcu.column_name ORDER BY kcu.ordinal_position) AS columns
        FROM information_schema.table_constraints AS tc
        JOIN information_schema.key_column_usage AS kcu
          ON kcu.constraint_schema = tc.constraint_schema
         AND kcu.constraint_name = tc.constraint_name
         AND kcu.table_schema = tc.table_schema
         AND kcu.table_name = tc.table_name
        WHERE tc.table_schema = ANY($1::text[])
          AND tc.constraint_type IN ('PRIMARY KEY', 'UNIQUE')
        GROUP BY tc.table_schema, tc.table_name, tc.constraint_name, tc.constraint_type
        ORDER BY
          tc.table_schema,
          tc.table_name,
          CASE tc.constraint_type WHEN 'UNIQUE' THEN 0 ELSE 1 END,
          tc.constraint_name
        """,
        list(ALLOWED_SCHEMAS),
    )
    unique_index_rows = await connection.fetch(
        """
        SELECT
            n.nspname AS table_schema,
            c.relname AS table_name,
            i.relname AS index_name,
            ARRAY_AGG(a.attname ORDER BY key_cols.ordinality) AS columns
        FROM pg_index AS ix
        JOIN pg_class AS c
          ON c.oid = ix.indrelid
        JOIN pg_namespace AS n
          ON n.oid = c.relnamespace
        JOIN pg_class AS i
          ON i.oid = ix.indexrelid
        JOIN unnest(ix.indkey) WITH ORDINALITY AS key_cols(attnum, ordinality)
          ON true
        JOIN pg_attribute AS a
          ON a.attrelid = c.oid
         AND a.attnum = key_cols.attnum
        JOIN information_schema.tables AS t
          ON t.table_schema = n.nspname
         AND t.table_name = c.relname
         AND t.table_type = 'BASE TABLE'
        WHERE n.nspname = ANY($1::text[])
          AND ix.indisunique
          AND ix.indisvalid
          AND ix.indpred IS NULL
        GROUP BY n.nspname, c.relname, i.relname
        ORDER BY n.nspname, c.relname, i.relname
        """,
        list(ALLOWED_SCHEMAS),
    )

    enum_rows = await connection.fetch(
        """
        SELECT t.typname, e.enumlabel
        FROM pg_type AS t
        JOIN pg_enum AS e
          ON e.enumtypid = t.oid
        ORDER BY t.typname, e.enumsortorder
        """
    )

    enum_labels_by_type: dict[str, list[str]] = {}
    for row in enum_rows:
        enum_labels_by_type.setdefault(row["typname"], []).append(row["enumlabel"])

    columns_by_table: dict[tuple[str, str], dict[str, ColumnMeta]] = {}
    for row in column_rows:
        table_key = (row["table_schema"], row["table_name"])
        columns_by_table.setdefault(table_key, {})[row["column_name"]] = ColumnMeta(
            name=row["column_name"],
            data_type=row["data_type"],
            udt_name=row["udt_name"],
            is_nullable=row["is_nullable"] == "YES",
            default=row["column_default"] or "",
            ordinal_position=row["ordinal_position"],
            is_identity=row["is_identity"] == "YES",
            enum_labels=tuple(enum_labels_by_type.get(row["udt_name"], ())),
        )

    conflict_candidates: dict[tuple[str, str], list[tuple[str, ...]]] = {}
    for row in [*constraint_rows, *unique_index_rows]:
        columns = tuple(row["columns"] or ())
        if columns:
            table_key = (row["table_schema"], row["table_name"])
            conflict_candidates.setdefault(table_key, []).append(columns)

    metadata: dict[str, TableMeta] = {}
    duplicate_table_names = {
        table_name
        for _, table_name in columns_by_table
        if sum(1 for _, candidate_name in columns_by_table if candidate_name == table_name) > 1
    }
    for (schema_name, table_name), columns in columns_by_table.items():
        if table_name in duplicate_table_names:
            continue
        conflict_columns, supports_on_conflict = resolve_table_conflict_columns(
            schema_name,
            table_name,
            conflict_candidates.get((schema_name, table_name), []),
            columns,
        )
        metadata[table_name] = TableMeta(
            schema_name=schema_name,
            name=table_name,
            columns=columns,
            conflict_columns=conflict_columns,
            supports_on_conflict=supports_on_conflict,
        )
    return metadata


def choose_conflict_columns(
    candidates: list[tuple[str, ...]],
    columns: dict[str, ColumnMeta],
) -> tuple[str, ...]:
    unique_candidates: list[tuple[str, ...]] = []
    for candidate in candidates:
        if candidate in unique_candidates:
            continue
        if any(column not in columns for column in candidate):
            continue
        unique_candidates.append(candidate)

    multi_column_candidates = [
        candidate for candidate in unique_candidates if len(candidate) > 1
    ]
    if multi_column_candidates:
        return multi_column_candidates[0]

    non_auto_candidates = [
        candidate
        for candidate in unique_candidates
        if not (len(candidate) == 1 and columns[candidate[0]].is_auto_generated)
    ]
    return non_auto_candidates[0] if non_auto_candidates else ()


def resolve_table_conflict_columns(
    schema_name: str,
    table_name: str,
    candidates: list[tuple[str, ...]],
    columns: dict[str, ColumnMeta],
) -> tuple[tuple[str, ...], bool]:
    hardcoded = LEDGER_CONFLICT_COLUMNS.get((schema_name, table_name))
    if hardcoded and all(column in columns for column in hardcoded):
        return hardcoded, False
    return choose_conflict_columns(candidates, columns), True


def is_backend_generated_column(table_meta: TableMeta, column_name: str) -> bool:
    return (
        (table_meta.schema_name, table_meta.name) in BACKEND_GENERATED_ID_TABLES
        and column_name == "id"
    )


def sheet_has_data(sheet: Any) -> bool:
    for row in sheet.iter_rows(values_only=True):
        if any(not is_blank(value) for value in row):
            return True
    return False


def get_header(sheet: Any) -> list[str]:
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return [normalize_header_value(value) for value in first_row]


def build_import_plan(content: bytes, metadata: dict[str, TableMeta]) -> list[PreparedSheet]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"无法读取 Excel 文件：{exc}") from exc

    sheets: list[PreparedSheet] = []
    for worksheet in workbook.worksheets:
        if not sheet_has_data(worksheet):
            continue

        sheet_result = PreparedSheet(
            sheet_name=worksheet.title,
            schema_name=None,
            table_name=worksheet.title,
        )
        table_meta = metadata.get(worksheet.title)
        if table_meta is None:
            sheet_result.errors.append(
                f"sheet 名称必须与 survey 或 ledger 中的可写表完全一致：{worksheet.title}"
            )
            sheets.append(sheet_result)
            continue

        sheet_result.schema_name = table_meta.schema_name
        if not table_meta.conflict_columns:
            sheet_result.errors.append("目标表缺少主键或唯一键，不能安全跳过重复记录")
            sheets.append(sheet_result)
            continue

        headers = get_header(worksheet)
        if validate_headers(headers, table_meta, sheet_result):
            sheets.append(sheet_result)
            continue

        header_indexes = {
            header: index
            for index, header in enumerate(headers)
            if header and header in table_meta.columns
        }
        ignored_auto_columns = {
            header
            for header in header_indexes
            if table_meta.columns[header].is_auto_generated
            or is_backend_generated_column(table_meta, header)
        }

        validate_required_columns(header_indexes, table_meta, sheet_result)
        if sheet_result.errors:
            sheets.append(sheet_result)
            continue

        parse_rows(worksheet, table_meta, header_indexes, ignored_auto_columns, sheet_result)
        sheets.append(sheet_result)

    return sheets


def validate_headers(
    headers: list[str],
    table_meta: TableMeta,
    sheet_result: PreparedSheet,
) -> bool:
    if not any(headers):
        sheet_result.errors.append("第 1 行必须是数据库列名")
        return True

    seen_headers: set[str] = set()
    for header in headers:
        if not header:
            continue
        if header in seen_headers:
            sheet_result.errors.append(f"列名重复：{header}")
        seen_headers.add(header)
        if header not in table_meta.columns:
            sheet_result.errors.append(f"目标表不存在列：{header}")

    return bool(sheet_result.errors)


def validate_required_columns(
    header_indexes: dict[str, int],
    table_meta: TableMeta,
    sheet_result: PreparedSheet,
) -> None:
    missing_columns = []
    for column in table_meta.columns.values():
        if (
            column.is_auto_generated
            or is_backend_generated_column(table_meta, column.name)
            or column.is_nullable
            or column.has_default
        ):
            continue
        if column.name not in header_indexes:
            missing_columns.append(column.name)

    if missing_columns:
        sheet_result.errors.append(f"缺少必填列：{', '.join(missing_columns)}")


def parse_rows(
    worksheet: Any,
    table_meta: TableMeta,
    header_indexes: dict[str, int],
    ignored_auto_columns: set[str],
    sheet_result: PreparedSheet,
) -> None:
    ordered_headers = [
        header
        for header, _ in sorted(header_indexes.items(), key=lambda item: item[1])
        if header not in ignored_auto_columns
    ]

    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(not is_blank(value) for value in row):
            continue

        sheet_result.row_count += 1
        row_values: dict[str, Any] = {}
        row_errors: list[str] = []
        for header in ordered_headers:
            column = table_meta.columns[header]
            index = header_indexes[header]
            raw_value = row[index] if index < len(row) else None
            try:
                parsed_value = parse_cell_value(column, raw_value)
            except ValueError as exc:
                row_errors.append(f"{header}：{exc}")
                continue

            if parsed_value is None and not column.is_nullable and not column.has_default:
                row_errors.append(f"{header} 不能为空")
                continue
            if parsed_value is USE_DEFAULT:
                continue
            row_values[header] = parsed_value

        missing_conflict_columns = [
            column
            for column in table_meta.conflict_columns
            if is_blank(row_values.get(column))
        ]
        if missing_conflict_columns:
            row_errors.append(f"冲突键字段不能为空：{', '.join(missing_conflict_columns)}")

        if row_errors:
            sheet_result.errors.append(f"第 {row_number} 行：{'；'.join(row_errors)}")
            continue

        conflict_values = tuple(row_values[column] for column in table_meta.conflict_columns)
        sheet_result.rows.append(
            PreparedRow(
                row_number=row_number,
                values=row_values,
                conflict_values=conflict_values,
            )
        )
        sheet_result.valid_rows += 1


def mark_file_duplicates(sheet_result: PreparedSheet) -> None:
    seen_keys: set[tuple[Any, ...]] = set()
    for row in sheet_result.rows:
        if row.conflict_values in seen_keys:
            row.skipped_duplicate = True
            sheet_result.skipped_duplicate_rows += 1
            sheet_result.warnings.append(
                f"第 {row.row_number} 行与本文件前面记录重复，已跳过"
            )
            continue
        seen_keys.add(row.conflict_values)


def parse_column_default_value(column: ColumnMeta) -> Any:
    """从列默认值表达式中提取字面量（如 2026、'第一代'::text），提取不到返回 None。"""
    text = column.default.strip()
    if not text:
        return None
    match = re.match(r"^'(.*)'::", text)
    if match:
        return match.group(1)
    if text.isdigit():
        return int(text)
    return None


def normalize_history_key_value(value: Any) -> Any:
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float)):
        return int(value)
    return str(value).strip()


def event_time_of(values: dict[str, Any]) -> datetime | None:
    value = values.get("事件时间")
    return value if isinstance(value, datetime) else None


async def correct_dispatch_event_types(
    connection: asyncpg.Connection,
    sheets: list[PreparedSheet],
    metadata: dict[str, TableMeta],
) -> None:
    """用户录入下派类事件但同组已存在历史事件时，把事件类型纠正为"复查异常"。

    历史 = 数据库已有事件 + 本文件中事件时间更早的行；分组键由
    LEDGER_HISTORY_RULES 定义（编号 + 年份/世代/虫害类型）。
    用户填写的防治、复查异常、复查合格一律信任，不做纠正。
    """
    for sheet in sheets:
        if sheet.table_name is None or sheet.table_name not in metadata:
            continue
        if sheet.errors:
            continue
        table_meta = metadata[sheet.table_name]
        rule = LEDGER_HISTORY_RULES.get((table_meta.schema_name, table_meta.name))
        if rule is None or not sheet.rows:
            continue
        group_columns, dispatch_types = rule

        codes = sorted(
            {
                str(row.values["编号"]).strip()
                for row in sheet.rows
                if not is_blank(row.values.get("编号"))
            }
        )
        if not codes:
            continue

        select_columns = list(dict.fromkeys([*group_columns, "事件类型", "事件时间"]))
        qualified_table = (
            f"{quote_identifier(table_meta.schema_name)}.{quote_identifier(table_meta.name)}"
        )
        history_rows = await connection.fetch(
            f"""
            SELECT {', '.join(quote_identifier(column) for column in select_columns)}
            FROM {qualified_table}
            WHERE {quote_identifier("编号")} = ANY($1::text[])
            """,
            codes,
        )

        def effective_group_key(values: dict[str, Any]) -> tuple[Any, ...] | None:
            key: list[Any] = []
            for column in group_columns:
                value = values.get(column)
                if is_blank(value):
                    value = parse_column_default_value(table_meta.columns[column])
                if is_blank(value) and column == "年份":
                    event_time = event_time_of(values)
                    if event_time is not None:
                        value = event_time.year
                if is_blank(value):
                    return None
                key.append(normalize_history_key_value(value))
            return tuple(key)

        histories: dict[tuple[Any, ...], list[tuple[str, Any]]] = {}
        for history_row in history_rows:
            event_type = history_row["事件类型"]
            if is_blank(event_type) or any(
                is_blank(history_row[column]) for column in group_columns
            ):
                continue
            key = tuple(
                normalize_history_key_value(history_row[column])
                for column in group_columns
            )
            histories.setdefault(key, []).append(
                (str(event_type).strip(), history_row["事件时间"])
            )

        ordered_rows = sorted(
            sheet.rows,
            key=lambda row: (
                event_time_of(row.values) is None,
                event_time_of(row.values) or datetime.min,
            ),
        )
        for row in ordered_rows:
            event_type = row.values.get("事件类型")
            key = effective_group_key(row.values)
            if key is None:
                continue
            current_time = event_time_of(row.values)
            if not is_blank(event_type) and str(event_type).strip() in dispatch_types:
                original = str(event_type).strip()
                qualifies = any(
                    h_type == original or h_type not in dispatch_types
                    for h_type, h_time in histories.get(key, [])
                    if h_time is None
                    or current_time is None
                    or h_time < current_time
                )
                if qualifies:
                    row.values["事件类型"] = RECHECK_ABNORMAL_EVENT_TYPE
                    row.conflict_values = tuple(
                        row.values[column] for column in table_meta.conflict_columns
                    )
                    sheet.warnings.append(
                        f"第 {row.row_number} 行：编号 {row.values.get('编号')} "
                        f"存在历史事件，事件类型由「{original}」纠正为"
                        f"「{RECHECK_ABNORMAL_EVENT_TYPE}」"
                    )
            if not is_blank(row.values.get("事件类型")):
                histories.setdefault(key, []).append(
                    (str(row.values["事件类型"]).strip(), current_time)
                )


def has_errors(sheets: list[PreparedSheet]) -> bool:
    return any(sheet.errors for sheet in sheets)


async def mark_database_duplicates(
    connection: asyncpg.Connection,
    sheets: list[PreparedSheet],
    metadata: dict[str, TableMeta],
) -> None:
    for sheet in sheets:
        if sheet.table_name is None or sheet.table_name not in metadata:
            continue
        if sheet.errors:
            continue

        table_meta = metadata[sheet.table_name]
        duplicate_keys: set[tuple[Any, ...]] = set()
        for row in sheet.rows:
            if row.skipped_duplicate or row.conflict_values in duplicate_keys:
                continue
            if await row_exists(connection, table_meta, row):
                row.skipped_duplicate = True
                duplicate_keys.add(row.conflict_values)
                sheet.skipped_duplicate_rows += 1


async def row_exists(
    connection: asyncpg.Connection,
    table_meta: TableMeta,
    row: PreparedRow,
) -> bool:
    qualified_table = (
        f"{quote_identifier(table_meta.schema_name)}.{quote_identifier(table_meta.name)}"
    )
    conditions = [
        f"{quote_identifier(column)} = ${index}"
        for index, column in enumerate(table_meta.conflict_columns, start=1)
    ]
    result = await connection.fetchrow(
        f"""
        SELECT 1
        FROM {qualified_table}
        WHERE {' AND '.join(conditions)}
        LIMIT 1
        """,
        *row.conflict_values,
    )
    return result is not None


async def insert_valid_rows(
    connection: asyncpg.Connection,
    sheets: list[PreparedSheet],
    metadata: dict[str, TableMeta],
) -> None:
    await assign_backend_generated_ids(connection, sheets, metadata)
    for sheet in sheets:
        if sheet.table_name is None or sheet.table_name not in metadata or sheet.errors:
            continue

        table_meta = metadata[sheet.table_name]
        for row in sheet.rows:
            if row.skipped_duplicate:
                continue
            inserted = await insert_row(connection, table_meta, row)
            if inserted:
                sheet.inserted_rows += 1
            else:
                row.skipped_duplicate = True
                sheet.skipped_duplicate_rows += 1


async def assign_backend_generated_ids(
    connection: asyncpg.Connection,
    sheets: list[PreparedSheet],
    metadata: dict[str, TableMeta],
) -> None:
    for sheet in sheets:
        if sheet.table_name is None or sheet.table_name not in metadata or sheet.errors:
            continue

        table_meta = metadata[sheet.table_name]
        if (table_meta.schema_name, table_meta.name) not in BACKEND_GENERATED_ID_TABLES:
            continue
        if "id" not in table_meta.columns:
            continue

        rows_requiring_id = [
            row for row in sheet.rows if not row.skipped_duplicate and "id" not in row.values
        ]
        if not rows_requiring_id:
            continue

        qualified_table = (
            f"{quote_identifier(table_meta.schema_name)}.{quote_identifier(table_meta.name)}"
        )
        await connection.execute(f"LOCK TABLE {qualified_table} IN SHARE ROW EXCLUSIVE MODE")
        next_id = (
            await connection.fetchval(f"SELECT COALESCE(MAX(id), 0) + 1 FROM {qualified_table}")
        ) or 1
        for row in rows_requiring_id:
            row.values["id"] = next_id
            next_id += 1


async def insert_row(
    connection: asyncpg.Connection,
    table_meta: TableMeta,
    row: PreparedRow,
) -> bool:
    ordered_columns = sorted(
        row.values,
        key=lambda name: table_meta.columns[name].ordinal_position,
    )
    placeholders = [f"${index}" for index in range(1, len(ordered_columns) + 1)]
    conflict_columns = ", ".join(quote_identifier(column) for column in table_meta.conflict_columns)
    qualified_table = (
        f"{quote_identifier(table_meta.schema_name)}.{quote_identifier(table_meta.name)}"
    )
    values = [row.values[column] for column in ordered_columns]
    if not table_meta.supports_on_conflict:
        conflict_conditions = [
            f"{quote_identifier(column)} = ${len(values) + index}"
            for index, column in enumerate(table_meta.conflict_columns, start=1)
        ]
        result = await connection.fetchrow(
            f"""
            INSERT INTO {qualified_table} (
                {', '.join(quote_identifier(column) for column in ordered_columns)}
            )
            SELECT {', '.join(placeholders)}
            WHERE NOT EXISTS (
                SELECT 1
                FROM {qualified_table}
                WHERE {' AND '.join(conflict_conditions)}
            )
            RETURNING 1 AS inserted
            """,
            *values,
            *row.conflict_values,
        )
        return result is not None

    result = await connection.fetchrow(
        f"""
        INSERT INTO {qualified_table} (
            {', '.join(quote_identifier(column) for column in ordered_columns)}
        )
        VALUES ({', '.join(placeholders)})
        ON CONFLICT ({conflict_columns}) DO NOTHING
        RETURNING 1 AS inserted
        """,
        *values,
    )
    return result is not None


def _normalize_group_label(value: Any, empty_label: str = EMPTY_LOCALITY_LABEL) -> str:
    if is_blank(value):
        return empty_label
    return str(value).strip()


def _row_damage_status(values: dict[str, Any]) -> bool | None:
    """Return True/False when damage can be judged, otherwise None."""
    if DAMAGED_PLANT_COUNT_FIELD in values:
        raw = values.get(DAMAGED_PLANT_COUNT_FIELD)
        try:
            count = 0 if raw is None or raw is USE_DEFAULT else int(raw)
        except (TypeError, ValueError):
            count = 0
        return count > 0

    if DAMAGE_LEVEL_FIELD in values:
        level = str(values.get(DAMAGE_LEVEL_FIELD) or "").strip()
        return level not in UNDAMAGED_LEVELS

    return None


def build_sheet_stats(sheet: PreparedSheet) -> dict[str, Any]:
    """Aggregate importable-row business stats for preview UI."""
    importable_rows = [
        row for row in sheet.rows if not row.skipped_duplicate
    ]
    locality_counts: dict[str, int] = {}
    event_type_counts: dict[str, int] = {}
    damaged_count = 0
    undamaged_count = 0
    has_locality = False
    has_event_type = False
    has_damage = False

    for row in importable_rows:
        values = row.values
        if LOCALITY_FIELD in values:
            has_locality = True
            label = _normalize_group_label(values.get(LOCALITY_FIELD))
            locality_counts[label] = locality_counts.get(label, 0) + 1

        if EVENT_TYPE_FIELD in values:
            has_event_type = True
            label = _normalize_group_label(
                values.get(EVENT_TYPE_FIELD),
                empty_label="未填写",
            )
            event_type_counts[label] = event_type_counts.get(label, 0) + 1

        damage_status = _row_damage_status(values)
        if damage_status is None:
            continue
        has_damage = True
        if damage_status:
            damaged_count += 1
        else:
            undamaged_count += 1

    by_locality = [
        {"name": name, "count": count}
        for name, count in sorted(
            locality_counts.items(),
            key=lambda item: (-item[1], item[0]),
        )
    ] if has_locality else []

    by_event_type = [
        {"name": name, "count": count}
        for name, count in sorted(
            event_type_counts.items(),
            key=lambda item: (-item[1], item[0]),
        )
    ] if has_event_type else []

    return {
        "by_locality": by_locality,
        "damaged_count": damaged_count if has_damage else None,
        "undamaged_count": undamaged_count if has_damage else None,
        "by_event_type": by_event_type,
    }


def summarize_import(
    *,
    file_name: str,
    dry_run: bool,
    sheets: list[PreparedSheet],
) -> dict[str, Any]:
    sheet_summaries = []
    for sheet in sheets:
        importable_rows = max(sheet.valid_rows - sheet.skipped_duplicate_rows, 0)
        sheet_summaries.append(
            {
                "sheet_name": sheet.sheet_name,
                "schema_name": sheet.schema_name,
                "table_name": sheet.table_name,
                "row_count": sheet.row_count,
                "valid_rows": sheet.valid_rows,
                "importable_rows": importable_rows,
                "inserted_rows": sheet.inserted_rows,
                "skipped_duplicate_rows": sheet.skipped_duplicate_rows,
                "warnings": sheet.warnings,
                "errors": sheet.errors,
                "stats": build_sheet_stats(sheet),
            }
        )

    totals = {
        "sheet_count": len(sheets),
        "row_count": sum(sheet.row_count for sheet in sheets),
        "valid_rows": sum(sheet.valid_rows for sheet in sheets),
        "importable_rows": sum(
            max(sheet.valid_rows - sheet.skipped_duplicate_rows, 0) for sheet in sheets
        ),
        "inserted_rows": sum(sheet.inserted_rows for sheet in sheets),
        "skipped_duplicate_rows": sum(sheet.skipped_duplicate_rows for sheet in sheets),
        "error_count": sum(len(sheet.errors) for sheet in sheets),
    }
    return {
        "file_name": file_name,
        "dry_run": dry_run,
        "totals": totals,
        "sheets": sheet_summaries,
    }


async def run_survey_excel_import(
    *,
    content: bytes,
    file_name: str,
    dry_run: bool,
    connection: asyncpg.Connection,
) -> dict[str, Any]:
    logger.info("开始 Excel 导入: file=%s size=%d bytes dry_run=%s", file_name, len(content), dry_run)
    metadata = await fetch_survey_table_metadata(connection)
    sheets = build_import_plan(content, metadata)

    sheet_summaries = [
        f"{sheet.sheet_name}(rows={sheet.row_count},valid={sheet.valid_rows},errors={len(sheet.errors)})"
        for sheet in sheets
    ]
    logger.info("Excel 解析完成: sheets=[%s]", ", ".join(sheet_summaries))

    if not sheets:
        return summarize_import(file_name=file_name, dry_run=dry_run, sheets=[])

    if not has_errors(sheets):
        await correct_dispatch_event_types(connection, sheets, metadata)
        for sheet in sheets:
            mark_file_duplicates(sheet)
        await mark_database_duplicates(connection, sheets, metadata)

    if dry_run or has_errors(sheets):
        summary = summarize_import(file_name=file_name, dry_run=dry_run, sheets=sheets)
        logger.info("Excel 预览完成: %s", summary.get("totals"))
        return summary

    async with connection.transaction():
        await insert_valid_rows(connection, sheets, metadata)

    summary = summarize_import(file_name=file_name, dry_run=dry_run, sheets=sheets)
    logger.info("Excel 入库完成: %s", summary.get("totals"))
    return summary


async def import_survey_excel(
    *,
    content: bytes,
    file_name: str,
    dry_run: bool,
) -> dict[str, Any]:
    pool = await ensure_pool()
    async with pool.acquire() as connection:
        return await run_survey_excel_import(
            content=content,
            file_name=file_name,
            dry_run=dry_run,
            connection=connection,
        )
