from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any

import asyncpg
from openpyxl import Workbook

from backend.db.postgres import ensure_pool, quote_identifier


ALLOWED_EXPORT_SCHEMAS = ("survey", "ledger")
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
INVALID_SHEET_NAME_CHARS = set('[]:*?/\\')
MAX_SHEET_NAME_LENGTH = 31

PEST_TABLE_MAPPING: dict[str, list[tuple[str, str]]] = {
    "美国白蛾": [
        ("survey", "美国白蛾调查表"),
        ("ledger", "美国白蛾问题点位事件流水表"),
        ("ledger", "美国白蛾问题点位台账"),
    ],
    "国槐尺蠖": [
        ("survey", "国槐尺蠖幼虫调查表"),
        ("ledger", "国槐尺蠖问题点位事件流水表"),
        ("ledger", "国槐尺蠖问题点位台账"),
    ],
    "春尺蠖": [
        ("survey", "春尺蠖成虫调查表"),
        ("survey", "春尺蠖幼虫调查表"),
        ("survey", "春尺蠖围环调查表"),
        ("ledger", "春尺蠖问题点位事件流水表"),
        ("ledger", "春尺蠖问题点位台账"),
    ],
    "其他害虫": [
        ("survey", "其他害虫调查表"),
        ("ledger", "其他害虫问题点位事件流水表"),
        ("ledger", "其他害虫问题点位台账"),
    ],
}


@dataclass(frozen=True)
class ExportTableMeta:
    schema_name: str
    table_name: str
    object_type: str
    columns: tuple[str, ...]
    row_count: int

    @property
    def column_count(self) -> int:
        return len(self.columns)

    def to_public_dict(self) -> dict[str, Any]:
        return {
            "schema_name": self.schema_name,
            "table_name": self.table_name,
            "object_type": self.object_type,
            "column_count": self.column_count,
            "row_count": self.row_count,
        }

    def to_summary_row(self) -> list[Any]:
        return [self.schema_name, self.table_name, self.object_type, self.row_count, self.column_count]


@dataclass(frozen=True)
class PestExportMeta:
    pest_type: str
    tables: tuple[ExportTableMeta, ...]
    total_row_count: int
    available_years: tuple[str, ...] = ()
    available_generations: tuple[str, ...] = ()

    def to_public_dict(self) -> dict[str, Any]:
        return {
            "pest_type": self.pest_type,
            "tables": [t.to_public_dict() for t in self.tables],
            "total_row_count": self.total_row_count,
            "available_years": list(self.available_years),
            "available_generations": list(self.available_generations),
        }


@dataclass(frozen=True)
class DataExportArtifact:
    filename: str
    media_type: str
    content: bytes


def validate_export_schema(schema_name: str) -> None:
    if schema_name not in ALLOWED_EXPORT_SCHEMAS:
        raise ValueError(f"不支持导出 schema：{schema_name}")


def normalize_sheet_name(value: str) -> str:
    normalized = "".join("_" if char in INVALID_SHEET_NAME_CHARS else char for char in value)
    normalized = normalized.strip().strip("'")
    return normalized or "导出表"


def build_unique_sheet_names(tables: list[ExportTableMeta]) -> dict[tuple[str, str], str]:
    used_names: set[str] = set()
    sheet_names: dict[tuple[str, str], str] = {}

    for table in tables:
        base_name = normalize_sheet_name(f"{table.schema_name}.{table.table_name}")
        candidate = base_name[:MAX_SHEET_NAME_LENGTH]
        suffix_index = 2
        while candidate in used_names:
            suffix = f"_{suffix_index}"
            candidate = f"{base_name[:MAX_SHEET_NAME_LENGTH - len(suffix)]}{suffix}"
            suffix_index += 1

        used_names.add(candidate)
        sheet_names[(table.schema_name, table.table_name)] = candidate

    return sheet_names


def serialize_cell_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool, date, datetime, Decimal)):
        return value
    if isinstance(value, (bytes, bytearray, memoryview)):
        return bytes(value).hex()
    if isinstance(value, (list, tuple, dict)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return str(value)


def build_export_filename(prefix: str, exported_at: datetime | None = None) -> str:
    timestamp = (exported_at or datetime.now()).strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{timestamp}.xlsx"


async def fetch_export_table_metadata(
    connection: asyncpg.Connection,
    schema_name: str | None = None,
    table_name: str | None = None,
) -> list[ExportTableMeta]:
    if schema_name is not None:
        validate_export_schema(schema_name)

    rows = await connection.fetch(
        """
        SELECT
            t.table_schema,
            t.table_name,
            CASE t.table_type
                WHEN 'VIEW' THEN 'view'
                ELSE 'table'
            END AS object_type,
            ARRAY_AGG(c.column_name ORDER BY c.ordinal_position) AS columns
        FROM information_schema.tables AS t
        JOIN information_schema.columns AS c
          ON c.table_schema = t.table_schema
         AND c.table_name = t.table_name
        WHERE t.table_schema = ANY($1::text[])
          AND t.table_type IN ('BASE TABLE', 'VIEW')
          AND ($2::text IS NULL OR t.table_schema = $2)
          AND ($3::text IS NULL OR t.table_name = $3)
        GROUP BY t.table_schema, t.table_name, t.table_type
        ORDER BY
            CASE t.table_schema WHEN 'survey' THEN 0 WHEN 'ledger' THEN 1 ELSE 2 END,
            CASE t.table_type WHEN 'BASE TABLE' THEN 0 WHEN 'VIEW' THEN 1 ELSE 2 END,
            t.table_name
        """,
        list(ALLOWED_EXPORT_SCHEMAS),
        schema_name,
        table_name,
    )

    tables: list[ExportTableMeta] = []
    for row in rows:
        qualified_table = (
            f"{quote_identifier(row['table_schema'])}.{quote_identifier(row['table_name'])}"
        )
        count_row = await connection.fetchrow(
            f"SELECT COUNT(*) AS row_count FROM {qualified_table}"
        )
        tables.append(
            ExportTableMeta(
                schema_name=row["table_schema"],
                table_name=row["table_name"],
                object_type=row["object_type"],
                columns=tuple(row["columns"] or ()),
                row_count=int(count_row["row_count"] if count_row else 0),
            )
        )

    return tables


async def list_export_tables() -> list[dict[str, Any]]:
    pool = await ensure_pool()
    async with pool.acquire() as connection:
        tables = await fetch_export_table_metadata(connection)
    return [table.to_public_dict() for table in tables]


async def get_export_table_meta(
    connection: asyncpg.Connection,
    schema_name: str,
    table_name: str,
) -> ExportTableMeta:
    tables = await fetch_export_table_metadata(
        connection,
        schema_name=schema_name,
        table_name=table_name,
    )
    if not tables:
        raise ValueError(f"表或视图不存在或不允许导出：{schema_name}.{table_name}")
    return tables[0]


async def _fetch_pest_filter_options(
    connection: asyncpg.Connection,
    pest_type: str,
    table_lookup: dict[tuple[str, str], ExportTableMeta],
) -> tuple[list[str], list[str]]:
    mapping = PEST_TABLE_MAPPING.get(pest_type, [])
    survey_tables = [(s, t) for s, t in mapping if s == "survey"]
    if not survey_tables:
        return [], []

    schema, table_name = survey_tables[0]
    meta = table_lookup.get((schema, table_name))
    if not meta:
        return [], []

    qualified = f"{quote_identifier(schema)}.{quote_identifier(table_name)}"
    has_generation = "世代" in meta.columns

    years: list[str] = []
    rows = await connection.fetch(
        f'SELECT DISTINCT "年份" FROM {qualified} WHERE "年份" IS NOT NULL ORDER BY "年份"'
    )
    years = [str(r["年份"]) for r in rows]

    generations: list[str] = []
    if has_generation:
        rows = await connection.fetch(
            f'SELECT DISTINCT "世代" FROM {qualified} WHERE "世代" IS NOT NULL ORDER BY "世代"'
        )
        generations = [str(r["世代"]) for r in rows]

    return years, generations


async def fetch_pest_export_metadata(
    connection: asyncpg.Connection,
    pest_type: str | None = None,
) -> list[PestExportMeta]:
    all_tables = await fetch_export_table_metadata(connection)
    table_lookup = {(t.schema_name, t.table_name): t for t in all_tables}

    pest_types_to_include = [pest_type] if pest_type else list(PEST_TABLE_MAPPING.keys())

    result: list[PestExportMeta] = []
    for pt in pest_types_to_include:
        if pt not in PEST_TABLE_MAPPING:
            raise ValueError(f"不支持的虫种：{pt}")

        tables: list[ExportTableMeta] = []
        for key in PEST_TABLE_MAPPING[pt]:
            table = table_lookup.get(key)
            if table:
                tables.append(table)

        total = sum(t.row_count for t in tables)
        years, generations = await _fetch_pest_filter_options(connection, pt, table_lookup)
        result.append(
            PestExportMeta(
                pest_type=pt,
                tables=tuple(tables),
                total_row_count=total,
                available_years=tuple(years),
                available_generations=tuple(generations),
            )
        )

    return result


async def list_pest_export_types() -> list[dict[str, Any]]:
    pool = await ensure_pool()
    async with pool.acquire() as connection:
        pest_metas = await fetch_pest_export_metadata(connection)
    return [pm.to_public_dict() for pm in pest_metas]


async def fetch_pest_export_metadata_filtered(
    connection: asyncpg.Connection,
    pest_type: str,
    year: str | None = None,
    generation: str | None = None,
) -> PestExportMeta:
    if pest_type not in PEST_TABLE_MAPPING:
        raise ValueError(f"不支持的虫种：{pest_type}")

    all_tables = await fetch_export_table_metadata(connection)
    table_lookup = {(t.schema_name, t.table_name): t for t in all_tables}

    tables: list[ExportTableMeta] = []
    for key in PEST_TABLE_MAPPING[pest_type]:
        table = table_lookup.get(key)
        if not table:
            continue

        conditions, params = _build_filter_params(table, year, generation)
        qualified_table = (
            f"{quote_identifier(table.schema_name)}.{quote_identifier(table.table_name)}"
        )

        if conditions:
            where_clause = " AND ".join(conditions)
            count_row = await connection.fetchrow(
                f"SELECT COUNT(*) AS row_count FROM {qualified_table} WHERE {where_clause}",
                *params,
            )
        else:
            count_row = await connection.fetchrow(
                f"SELECT COUNT(*) AS row_count FROM {qualified_table}"
            )

        filtered_count = int(count_row["row_count"] if count_row else 0)
        tables.append(
            ExportTableMeta(
                schema_name=table.schema_name,
                table_name=table.table_name,
                object_type=table.object_type,
                columns=table.columns,
                row_count=filtered_count,
            )
        )

    total = sum(t.row_count for t in tables)
    years, generations = await _fetch_pest_filter_options(connection, pest_type, table_lookup)

    return PestExportMeta(
        pest_type=pest_type,
        tables=tuple(tables),
        total_row_count=total,
        available_years=tuple(years),
        available_generations=tuple(generations),
    )


def _build_filter_params(
    table: ExportTableMeta,
    year: str | None,
    generation: str | None,
) -> tuple[list[str], list[str]]:
    conditions: list[str] = []
    params: list[str] = []
    columns_set = set(table.columns)
    idx = 0

    if year is not None and "年份" in columns_set:
        idx += 1
        conditions.append(f'"年份" = ${idx}::text')
        params.append(year)

    if generation is not None and "世代" in columns_set:
        idx += 1
        conditions.append(f'"世代" = ${idx}::text')
        params.append(generation)

    return conditions, params


async def append_table_sheet(
    workbook: Workbook,
    connection: asyncpg.Connection,
    table: ExportTableMeta,
    sheet_name: str,
    where_clause: str | None = None,
    where_params: tuple[Any, ...] = (),
) -> None:
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.append(list(table.columns))

    qualified_table = f"{quote_identifier(table.schema_name)}.{quote_identifier(table.table_name)}"
    column_sql = ", ".join(quote_identifier(column) for column in table.columns)

    if where_clause:
        rows = await connection.fetch(
            f"SELECT {column_sql} FROM {qualified_table} WHERE {where_clause}",
            *where_params,
        )
    else:
        rows = await connection.fetch(f"SELECT {column_sql} FROM {qualified_table}")

    for row in rows:
        worksheet.append([serialize_cell_value(row[column]) for column in table.columns])


def append_summary_sheet(
    workbook: Workbook,
    tables: list[ExportTableMeta],
    sheet_names: dict[tuple[str, str], str],
    exported_at: datetime,
    range_description: str = "survey, ledger 表和视图",
) -> None:
    worksheet = workbook.create_sheet("导出说明", 0)
    worksheet.append(["导出时间", exported_at.strftime("%Y-%m-%d %H:%M:%S")])
    worksheet.append(["导出范围", range_description])
    worksheet.append([])
    worksheet.append(["schema", "名称", "类型", "sheet 名", "记录数", "字段数"])
    for table in tables:
        worksheet.append(
            [
                table.schema_name,
                table.table_name,
                "视图" if table.object_type == "view" else "表",
                sheet_names[(table.schema_name, table.table_name)],
                table.row_count,
                table.column_count,
            ]
        )


def workbook_to_bytes(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


async def export_all_tables() -> DataExportArtifact:
    exported_at = datetime.now()
    workbook = Workbook(write_only=True)

    pool = await ensure_pool()
    async with pool.acquire() as connection:
        tables = await fetch_export_table_metadata(connection)
        if not tables:
            raise ValueError("没有可导出的 survey 或 ledger 表或视图")

        sheet_names = build_unique_sheet_names(tables)
        append_summary_sheet(workbook, tables, sheet_names, exported_at)
        for table in tables:
            await append_table_sheet(
                workbook,
                connection,
                table,
                sheet_names[(table.schema_name, table.table_name)],
            )

    return DataExportArtifact(
        filename=build_export_filename("调查数据导出", exported_at),
        media_type=XLSX_MEDIA_TYPE,
        content=workbook_to_bytes(workbook),
    )


async def export_single_table(schema_name: str, table_name: str) -> DataExportArtifact:
    exported_at = datetime.now()
    workbook = Workbook(write_only=True)

    pool = await ensure_pool()
    async with pool.acquire() as connection:
        table = await get_export_table_meta(connection, schema_name, table_name)
        sheet_name = build_unique_sheet_names([table])[(table.schema_name, table.table_name)]
        await append_table_sheet(workbook, connection, table, sheet_name)

    return DataExportArtifact(
        filename=build_export_filename(f"{schema_name}_{table_name}", exported_at),
        media_type=XLSX_MEDIA_TYPE,
        content=workbook_to_bytes(workbook),
    )


async def export_pest_type(
    pest_type: str,
    year: str | None = None,
    generation: str | None = None,
) -> DataExportArtifact:
    exported_at = datetime.now()
    workbook = Workbook(write_only=True)

    pool = await ensure_pool()
    async with pool.acquire() as connection:
        pest_metas = await fetch_pest_export_metadata(connection, pest_type=pest_type)
        if not pest_metas or not pest_metas[0].tables:
            raise ValueError(f"虫种不存在或无数据：{pest_type}")

        tables = list(pest_metas[0].tables)
        sheet_names = build_unique_sheet_names(tables)

        filter_parts = [f"虫种：{pest_type}"]
        if year:
            filter_parts.append(f"年份={year}")
        if generation:
            filter_parts.append(f"世代={generation}")

        append_summary_sheet(
            workbook,
            tables,
            sheet_names,
            exported_at,
            range_description="，".join(filter_parts),
        )

        for table in tables:
            conditions, params = _build_filter_params(table, year, generation)
            where_clause = " AND ".join(conditions) if conditions else None
            await append_table_sheet(
                workbook,
                connection,
                table,
                sheet_names[(table.schema_name, table.table_name)],
                where_clause=where_clause,
                where_params=tuple(params),
            )

    filter_suffix = ""
    if year or generation:
        parts = []
        if year:
            parts.append(f"年份{year}")
        if generation:
            parts.append(f"世代{generation}")
        filter_suffix = f"_{'-'.join(parts)}"

    return DataExportArtifact(
        filename=build_export_filename(f"{pest_type}{filter_suffix}", exported_at),
        media_type=XLSX_MEDIA_TYPE,
        content=workbook_to_bytes(workbook),
    )
