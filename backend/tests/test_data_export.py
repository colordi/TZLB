from __future__ import annotations

import unittest
from datetime import date
from io import BytesIO
from unittest.mock import AsyncMock, patch

from openpyxl import Workbook, load_workbook

from backend.routers.data_export import build_download_response, get_pest_export_meta
from backend.services.data_export import (
    DataExportArtifact,
    ExportTableMeta,
    append_table_sheet,
    build_unique_sheet_names,
    export_all_tables,
    export_pest_type,
    fetch_export_table_metadata,
    fetch_pest_export_metadata,
    fetch_pest_export_metadata_filtered,
    normalize_date_list_text,
    workbook_to_bytes,
)


class FakeAcquire:
    def __init__(self, connection):
        self.connection = connection

    async def __aenter__(self):
        return self.connection

    async def __aexit__(self, exc_type, exc, traceback):
        return False


class FakePool:
    def __init__(self, connection):
        self.connection = connection

    def acquire(self):
        return FakeAcquire(self.connection)


class FakeConnection:
    def __init__(self):
        self.fetch_calls: list[tuple[str, tuple]] = []
        self.fetchrow_calls: list[str] = []
        self.metadata_rows = [
            {
                "table_schema": "survey",
                "table_name": "春尺蠖幼虫调查表",
                "object_type": "table",
                "columns": ["编号", "调查日期"],
            },
            {
                "table_schema": "ledger",
                "table_name": "美国白蛾问题点位台账",
                "object_type": "table",
                "columns": ["编号", "备注"],
            },
            {
                "table_schema": "ledger",
                "table_name": "美国白蛾问题点位视图",
                "object_type": "view",
                "columns": ["编号", "属地"],
            },
        ]
        self.counts = {
            '"survey"."春尺蠖幼虫调查表"': 1,
            '"ledger"."美国白蛾问题点位台账"': 0,
            '"ledger"."美国白蛾问题点位视图"': 1,
        }
        self.table_rows = {
            '"survey"."春尺蠖幼虫调查表"': [
                {
                    "编号": "YF001",
                    "调查日期": date(2026, 4, 1),
                }
            ],
            '"ledger"."美国白蛾问题点位台账"': [],
            '"ledger"."美国白蛾问题点位视图"': [
                {
                    "编号": "MQ001",
                    "属地": "马驹桥镇",
                }
            ],
        }

    async def fetch(self, query: str, *args):
        self.fetch_calls.append((query, args))
        if "information_schema.tables" in query:
            schema_name = args[1]
            table_name = args[2]
            return [
                row
                for row in self.metadata_rows
                if (schema_name is None or row["table_schema"] == schema_name)
                and (table_name is None or row["table_name"] == table_name)
            ]

        for qualified_table, rows in self.table_rows.items():
            if qualified_table in query:
                return rows
        return []

    async def fetchrow(self, query: str, *args):
        self.fetchrow_calls.append(query)
        for qualified_table, count in self.counts.items():
            if qualified_table in query:
                return {"row_count": count}
        return {"row_count": 0}


class DataExportServiceTest(unittest.IsolatedAsyncioTestCase):
    async def test_metadata_reads_allowed_survey_and_ledger_tables_and_views(self) -> None:
        connection = FakeConnection()

        tables = await fetch_export_table_metadata(connection)

        self.assertEqual([table.schema_name for table in tables], ["survey", "ledger", "ledger"])
        self.assertEqual(tables[0].table_name, "春尺蠖幼虫调查表")
        self.assertEqual(tables[0].column_count, 2)
        self.assertEqual(tables[0].row_count, 1)
        self.assertEqual(tables[2].object_type, "view")
        self.assertEqual(connection.fetch_calls[0][1][0], ["survey", "ledger"])

    async def test_metadata_rejects_unknown_schema(self) -> None:
        with self.assertRaises(ValueError) as context:
            await fetch_export_table_metadata(FakeConnection(), schema_name="sites")

        self.assertEqual(str(context.exception), "不支持导出 schema：sites")

    async def test_export_all_tables_contains_summary_and_table_sheets(self) -> None:
        connection = FakeConnection()

        with patch(
            "backend.services.data_export.service.ensure_pool",
            new=AsyncMock(return_value=FakePool(connection)),
        ):
            artifact = await export_all_tables()

        workbook = load_workbook(BytesIO(artifact.content), read_only=True, data_only=True)
        self.assertIn("导出说明", workbook.sheetnames)
        self.assertIn("survey.春尺蠖幼虫调查表", workbook.sheetnames)
        self.assertIn("ledger.美国白蛾问题点位台账", workbook.sheetnames)
        self.assertIn("ledger.美国白蛾问题点位视图", workbook.sheetnames)

        summary_rows = list(workbook["导出说明"].iter_rows(values_only=True))
        self.assertIn(("导出范围", "survey, ledger 表和视图"), summary_rows)
        self.assertIn(
            (
                "survey",
                "春尺蠖幼虫调查表",
                "表",
                "survey.春尺蠖幼虫调查表",
                1,
                2,
            ),
            summary_rows,
        )
        survey_rows = list(workbook["survey.春尺蠖幼虫调查表"].iter_rows(values_only=True))
        self.assertEqual(survey_rows[0], ("编号", "调查日期"))
        self.assertEqual(survey_rows[1][0], "YF001")

    def test_sheet_names_are_truncated_and_deduplicated(self) -> None:
        sheet_names = build_unique_sheet_names(
            [
                ExportTableMeta(
                    schema_name="survey",
                    table_name="一个非常非常非常非常非常非常非常非常长的调查表名称",
                    object_type="table",
                    columns=("编号",),
                    row_count=0,
                ),
                ExportTableMeta(
                    schema_name="ledger",
                    table_name="一个非常非常非常非常非常非常非常非常长的调查表名称",
                    object_type="view",
                    columns=("编号",),
                    row_count=0,
                ),
            ]
        )

        names = list(sheet_names.values())
        self.assertEqual(len(names), 2)
        self.assertEqual(len(names[0]), 31)
        self.assertEqual(len(names[1]), 31)
        self.assertNotEqual(names[0], names[1])


def _build_rich_fake_connection():
    conn = FakeConnection()
    conn.metadata_rows = [
        {
            "table_schema": "survey",
            "table_name": "春尺蠖幼虫调查表",
            "object_type": "table",
            "columns": ["编号", "调查日期", "年份"],
        },
        {
            "table_schema": "survey",
            "table_name": "美国白蛾调查表",
            "object_type": "table",
            "columns": ["编号", "受害株数", "年份", "世代"],
        },
        {
            "table_schema": "survey",
            "table_name": "国槐尺蠖幼虫调查表",
            "object_type": "table",
            "columns": ["编号", "危害程度", "年份", "世代"],
        },
        {
            "table_schema": "survey",
            "table_name": "春尺蠖成虫调查表",
            "object_type": "table",
            "columns": ["编号", "年份"],
        },
        {
            "table_schema": "survey",
            "table_name": "春尺蠖围环调查表",
            "object_type": "table",
            "columns": ["编号", "年份"],
        },
        {
            "table_schema": "survey",
            "table_name": "其他害虫调查表",
            "object_type": "table",
            "columns": ["编号", "年份"],
        },
        {
            "table_schema": "survey",
            "table_name": "白蜡蛀干害虫调查表",
            "object_type": "table",
            "columns": [
                "编号",
                "属地",
                "调查日期",
                "目测死亡（株）",
                "伐除（株）",
                "换植（株）",
                "窄吉丁危害（株）",
                "窄吉丁孔数（个）",
                "木蠹蛾危害（株）",
                "备注",
                "年份",
            ],
        },
        {
            "table_schema": "ledger",
            "table_name": "美国白蛾问题点位事件流水表",
            "object_type": "table",
            "columns": ["编号", "事件类型", "年份", "世代"],
        },
        {
            "table_schema": "ledger",
            "table_name": "美国白蛾问题点位台账",
            "object_type": "view",
            "columns": ["编号", "当前状态", "年份", "世代"],
        },
        {
            "table_schema": "ledger",
            "table_name": "国槐尺蠖问题点位事件流水表",
            "object_type": "table",
            "columns": ["编号", "年份"],
        },
        {
            "table_schema": "ledger",
            "table_name": "国槐尺蠖问题点位台账",
            "object_type": "view",
            "columns": ["编号", "年份"],
        },
        {
            "table_schema": "ledger",
            "table_name": "春尺蠖问题点位事件流水表",
            "object_type": "table",
            "columns": ["编号", "年份"],
        },
        {
            "table_schema": "ledger",
            "table_name": "春尺蠖问题点位台账",
            "object_type": "view",
            "columns": ["编号", "年份"],
        },
        {
            "table_schema": "survey",
            "table_name": "杨树食叶害虫调查表",
            "object_type": "table",
            "columns": ["编号", "虫害类型", "年份"],
        },
        {
            "table_schema": "ledger",
            "table_name": "杨树食叶害虫问题点位事件流水表",
            "object_type": "table",
            "columns": ["编号", "年份"],
        },
        {
            "table_schema": "ledger",
            "table_name": "杨树食叶害虫问题点位台账",
            "object_type": "view",
            "columns": ["编号", "年份"],
        },
        {
            "table_schema": "ledger",
            "table_name": "其他害虫问题点位事件流水表",
            "object_type": "table",
            "columns": ["编号", "年份"],
        },
        {
            "table_schema": "ledger",
            "table_name": "其他害虫问题点位台账",
            "object_type": "view",
            "columns": ["编号", "年份"],
        },
    ]
    conn.counts = {
        '"survey"."春尺蠖幼虫调查表"': 30,
        '"survey"."美国白蛾调查表"': 50,
        '"survey"."国槐尺蠖幼虫调查表"': 20,
        '"survey"."春尺蠖成虫调查表"': 10,
        '"survey"."春尺蠖围环调查表"': 5,
        '"survey"."其他害虫调查表"': 8,
        '"survey"."白蜡蛀干害虫调查表"': 0,
        '"ledger"."美国白蛾问题点位事件流水表"': 40,
        '"ledger"."美国白蛾问题点位台账"': 25,
        '"ledger"."国槐尺蠖问题点位事件流水表"': 15,
        '"ledger"."国槐尺蠖问题点位台账"': 10,
        '"ledger"."春尺蠖问题点位事件流水表"': 12,
        '"ledger"."春尺蠖问题点位台账"': 8,
        '"ledger"."其他害虫问题点位事件流水表"': 6,
        '"ledger"."其他害虫问题点位台账"': 4,
        '"survey"."杨树食叶害虫调查表"': 7,
        '"ledger"."杨树食叶害虫问题点位事件流水表"': 3,
        '"ledger"."杨树食叶害虫问题点位台账"': 2,
    }
    conn.table_rows = {
        '"survey"."春尺蠖幼虫调查表"': [
            {"编号": "YF001", "调查日期": date(2026, 4, 1), "年份": "2026"},
            {"编号": "YF002", "调查日期": date(2026, 4, 2), "年份": "2025"},
        ],
        '"survey"."美国白蛾调查表"': [
            {"编号": "MB001", "受害株数": 5, "年份": "2026", "世代": "1"},
            {"编号": "MB002", "受害株数": 3, "年份": "2026", "世代": "2"},
            {"编号": "MB003", "受害株数": 2, "年份": "2025", "世代": "1"},
        ],
        '"survey"."国槐尺蠖幼虫调查表"': [
            {"编号": "GH001", "危害程度": "中", "年份": "2026", "世代": "1"},
        ],
        '"survey"."春尺蠖成虫调查表"': [
            {"编号": "CC001", "年份": "2026"},
        ],
        '"survey"."春尺蠖围环调查表"': [
            {"编号": "CC001", "年份": "2026"},
        ],
        '"survey"."其他害虫调查表"': [
            {"编号": "QT001", "年份": "2026"},
        ],
        '"survey"."白蜡蛀干害虫调查表"': [],
    }
    return conn


class PestExportServiceTest(unittest.IsolatedAsyncioTestCase):
    async def test_fetch_pest_export_metadata_returns_all_pest_types(self) -> None:
        connection = _build_rich_fake_connection()
        result = await fetch_pest_export_metadata(connection)
        pest_types = [pm.pest_type for pm in result]
        self.assertEqual(
            pest_types,
            ["美国白蛾", "国槐尺蠖", "春尺蠖", "其他害虫", "杨树食叶害虫", "白蜡蛀干害虫"],
        )
        yangshu_meta = next(pm for pm in result if pm.pest_type == "杨树食叶害虫")
        self.assertEqual(len(yangshu_meta.tables), 3)

    async def test_fetch_pest_export_metadata_filters_by_pest_type(self) -> None:
        connection = _build_rich_fake_connection()
        result = await fetch_pest_export_metadata(connection, pest_type="春尺蠖")
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0].pest_type, "春尺蠖")
        self.assertEqual(len(result[0].tables), 5)

    async def test_fetch_pest_export_metadata_rejects_unknown_pest(self) -> None:
        connection = _build_rich_fake_connection()
        with self.assertRaises(ValueError) as context:
            await fetch_pest_export_metadata(connection, pest_type="未知虫种")
        self.assertEqual(str(context.exception), "不支持的虫种：未知虫种")

    async def test_export_pest_type_creates_xlsx_with_correct_sheets(self) -> None:
        connection = _build_rich_fake_connection()
        with patch(
            "backend.services.data_export.service.ensure_pool",
            new=AsyncMock(return_value=FakePool(connection)),
        ):
            artifact = await export_pest_type("美国白蛾")

        workbook = load_workbook(BytesIO(artifact.content), read_only=True, data_only=True)
        self.assertIn("导出说明", workbook.sheetnames)
        self.assertIn("survey.美国白蛾调查表", workbook.sheetnames)
        self.assertIn("ledger.美国白蛾问题点位事件流水表", workbook.sheetnames)
        self.assertIn("ledger.美国白蛾问题点位台账", workbook.sheetnames)

        summary_rows = list(workbook["导出说明"].iter_rows(values_only=True))
        self.assertIn(("导出范围", "虫种：美国白蛾"), summary_rows)
        self.assertEqual(len(workbook.sheetnames), 4)

    async def test_export_pest_type_with_year_filter_builds_correct_where_clause(self) -> None:
        connection = _build_rich_fake_connection()
        original_fetch = connection.fetch

        async def capturing_fetch(query, *args):
            connection._last_query = query
            connection._last_args = args
            return await original_fetch(query, *args)

        connection.fetch = capturing_fetch

        with patch(
            "backend.services.data_export.service.ensure_pool",
            new=AsyncMock(return_value=FakePool(connection)),
        ):
            artifact = await export_pest_type("美国白蛾", year="2026")

        workbook = load_workbook(BytesIO(artifact.content), read_only=True, data_only=True)

        summary_rows = list(workbook["导出说明"].iter_rows(values_only=True))
        self.assertIn(("导出范围", "虫种：美国白蛾，年份=2026"), summary_rows)

    async def test_export_pest_type_with_year_and_generation_filter(self) -> None:
        connection = _build_rich_fake_connection()
        with patch(
            "backend.services.data_export.service.ensure_pool",
            new=AsyncMock(return_value=FakePool(connection)),
        ):
            artifact = await export_pest_type("美国白蛾", year="2026", generation="1")

        workbook = load_workbook(BytesIO(artifact.content), read_only=True, data_only=True)

        summary_rows = list(workbook["导出说明"].iter_rows(values_only=True))
        self.assertIn(("导出范围", "虫种：美国白蛾，年份=2026，世代=1"), summary_rows)

    async def test_fetch_pest_export_metadata_includes_filter_options(self) -> None:
        connection = _build_rich_fake_connection()
        result = await fetch_pest_export_metadata(connection, pest_type="美国白蛾")
        self.assertEqual(len(result), 1)
        pm = result[0]
        self.assertIn("2026", pm.available_years)
        self.assertIn("2025", pm.available_years)
        self.assertIn("1", pm.available_generations)
        self.assertIn("2", pm.available_generations)

    async def test_fetch_pest_export_metadata_filter_options_no_generation(self) -> None:
        connection = _build_rich_fake_connection()
        result = await fetch_pest_export_metadata(connection, pest_type="春尺蠖")
        pm = result[0]
        self.assertIn("2026", pm.available_years)
        self.assertEqual(len(pm.available_generations), 0)

    async def test_fetch_pest_export_metadata_filtered_applies_year_and_generation(self) -> None:
        connection = _build_rich_fake_connection()
        pm = await fetch_pest_export_metadata_filtered(
            connection, pest_type="美国白蛾", year="2026", generation="1"
        )

        self.assertEqual(pm.pest_type, "美国白蛾")
        self.assertGreater(len(pm.tables), 0)
        self.assertIn("2026", pm.available_years)
        self.assertIn("1", pm.available_generations)

        # 验证按条件过滤后执行了带 WHERE 的 COUNT 查询
        count_queries = [q for q in connection.fetchrow_calls if "COUNT(*)" in q]
        self.assertTrue(any('"年份"::text = $1' in q for q in count_queries))
        self.assertTrue(any('"世代" = $2::text' in q for q in count_queries))

    async def test_fetch_pest_export_metadata_filtered_unknown_pest_raises(self) -> None:
        connection = _build_rich_fake_connection()
        with self.assertRaises(ValueError) as context:
            await fetch_pest_export_metadata_filtered(connection, pest_type="未知虫种")
        self.assertEqual(str(context.exception), "不支持的虫种：未知虫种")

    async def test_export_pest_type_with_no_tables_raises_error(self) -> None:
        connection = FakeConnection()
        connection.metadata_rows = []
        connection.counts = {}
        connection.table_rows = {}
        with patch(
            "backend.services.data_export.service.ensure_pool",
            new=AsyncMock(return_value=FakePool(connection)),
        ):
            with self.assertRaises(ValueError) as context:
                await export_pest_type("美国白蛾")
            self.assertIn("虫种不存在或无数据", str(context.exception))


class DataExportRouterTest(unittest.IsolatedAsyncioTestCase):
    def test_download_response_sets_xlsx_attachment_headers(self) -> None:
        response = build_download_response(
            DataExportArtifact(
                filename="调查数据导出_20260606_120000.xlsx",
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                content=b"xlsx-bytes",
            )
        )

        self.assertEqual(response.body, b"xlsx-bytes")
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response.headers["content-type"],
        )
        self.assertIn("attachment;", response.headers["content-disposition"])
        self.assertIn(
            "%E8%B0%83%E6%9F%A5%E6%95%B0%E6%8D%AE%E5%AF%BC%E5%87%BA_20260606_120000.xlsx",
            response.headers["content-disposition"],
        )

    async def test_get_pest_export_meta_router_returns_filtered_meta(self) -> None:
        connection = _build_rich_fake_connection()
        with patch(
            "backend.routers.data_export.ensure_pool",
            new=AsyncMock(return_value=FakePool(connection)),
        ):
            result = await get_pest_export_meta("美国白蛾", year="2026", generation="1")

        self.assertEqual(result["pest_type"], "美国白蛾")
        self.assertIn("tables", result)
        self.assertIn("total_row_count", result)


class DateListColumnExportTest(unittest.IsolatedAsyncioTestCase):
    def test_normalize_date_list_text_pads_month_and_day(self) -> None:
        self.assertEqual(normalize_date_list_text("2026/7/6"), "2026-07-06")
        self.assertEqual(normalize_date_list_text("2026/7/6、2026/12/18"), "2026-07-06、2026-12-18")
        self.assertEqual(normalize_date_list_text("2026/11/30"), "2026-11-30")

    def test_normalize_date_list_text_keeps_other_values(self) -> None:
        self.assertEqual(normalize_date_list_text(""), "")
        self.assertEqual(normalize_date_list_text("待定"), "待定")
        self.assertIsNone(normalize_date_list_text(None))

    async def test_append_table_sheet_formats_date_list_columns_as_text(self) -> None:
        connection = FakeConnection()
        connection.table_rows = {
            '"ledger"."美国白蛾问题点位台账"': [
                {"编号": "MB001", "调查日期列表": "2026/7/6、2026/7/18", "备注": "复查"},
            ],
        }
        table = ExportTableMeta(
            schema_name="ledger",
            table_name="美国白蛾问题点位台账",
            object_type="view",
            columns=("编号", "调查日期列表", "备注"),
            row_count=1,
        )

        workbook = Workbook(write_only=True)
        await append_table_sheet(workbook, connection, table, "台账")

        loaded = load_workbook(BytesIO(workbook_to_bytes(workbook)))
        sheet = loaded["台账"]
        self.assertEqual(sheet.cell(row=1, column=2).value, "调查日期列表")
        self.assertEqual(sheet.cell(row=2, column=1).value, "MB001")
        self.assertEqual(sheet.cell(row=2, column=2).value, "2026-07-06、2026-07-18")
        self.assertEqual(sheet.cell(row=2, column=2).number_format, "@")
        self.assertEqual(sheet.cell(row=2, column=3).value, "复查")
        self.assertEqual(sheet.cell(row=2, column=3).number_format, "General")


if __name__ == "__main__":
    unittest.main()
