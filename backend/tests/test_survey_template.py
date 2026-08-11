from __future__ import annotations

import unittest
from io import BytesIO

from openpyxl import load_workbook

from backend.services.survey_excel_import import ColumnMeta, TableMeta
from backend.services.survey_template import (
    build_import_template_bytes,
    filter_template_metadata,
)


class SurveyTemplateTest(unittest.TestCase):
    def test_template_contains_columns_and_example_row(self) -> None:
        metadata = {
            "春尺蠖幼虫调查表": TableMeta(
                schema_name="survey",
                name="春尺蠖幼虫调查表",
                columns={
                    "编号": ColumnMeta(
                        name="编号",
                        data_type="character varying",
                        udt_name="varchar",
                        is_nullable=False,
                        default="",
                        ordinal_position=1,
                    ),
                    "调查日期": ColumnMeta(
                        name="调查日期",
                        data_type="date",
                        udt_name="date",
                        is_nullable=False,
                        default="",
                        ordinal_position=2,
                    ),
                    "受害株数": ColumnMeta(
                        name="受害株数",
                        data_type="integer",
                        udt_name="int4",
                        is_nullable=True,
                        default="0",
                        ordinal_position=3,
                    ),
                    "备注": ColumnMeta(
                        name="备注",
                        data_type="text",
                        udt_name="text",
                        is_nullable=True,
                        default="",
                        ordinal_position=4,
                    ),
                },
                conflict_columns=("编号",),
            ),
        }

        content = build_import_template_bytes(metadata)
        workbook = load_workbook(BytesIO(content))

        self.assertEqual(workbook.sheetnames, ["春尺蠖幼虫调查表"])
        worksheet = workbook.active
        self.assertIsNotNone(worksheet)

        headers = [cell.value for cell in worksheet[1]]
        self.assertEqual(headers, ["编号", "调查日期", "受害株数", "备注"])

        examples = [cell.value for cell in worksheet[2]]
        self.assertEqual(examples[0], "示例值")
        self.assertEqual(examples[1], "2026-05-01")
        self.assertEqual(examples[2], 1)
        self.assertEqual(examples[3], "示例值")

    def test_required_columns_are_marked(self) -> None:
        metadata = {
            "测试表": TableMeta(
                schema_name="survey",
                name="测试表",
                columns={
                    "必填列": ColumnMeta(
                        name="必填列",
                        data_type="character varying",
                        udt_name="varchar",
                        is_nullable=False,
                        default="",
                        ordinal_position=1,
                    ),
                    "可空列": ColumnMeta(
                        name="可空列",
                        data_type="text",
                        udt_name="text",
                        is_nullable=True,
                        default="",
                        ordinal_position=2,
                    ),
                },
                conflict_columns=("必填列",),
            ),
        }

        content = build_import_template_bytes(metadata)
        workbook = load_workbook(BytesIO(content))
        worksheet = workbook.active
        self.assertIsNotNone(worksheet)

        required_header = worksheet.cell(row=1, column=1)
        optional_header = worksheet.cell(row=1, column=2)
        self.assertTrue(required_header.font.bold)
        self.assertIsNotNone(required_header.comment)
        self.assertIn("必填", required_header.comment.text)
        self.assertNotEqual(required_header.fill.start_color.rgb, optional_header.fill.start_color.rgb)


    def test_example_row_prefers_real_data(self) -> None:
        metadata = {
            "美国白蛾调查表": TableMeta(
                schema_name="survey",
                name="美国白蛾调查表",
                columns={
                    "编号": ColumnMeta(
                        name="编号",
                        data_type="character varying",
                        udt_name="varchar",
                        is_nullable=False,
                        default="",
                        ordinal_position=1,
                    ),
                    "受害株数": ColumnMeta(
                        name="受害株数",
                        data_type="integer",
                        udt_name="int4",
                        is_nullable=True,
                        default="0",
                        ordinal_position=2,
                    ),
                    "备注": ColumnMeta(
                        name="备注",
                        data_type="text",
                        udt_name="text",
                        is_nullable=True,
                        default="",
                        ordinal_position=3,
                    ),
                    "id": ColumnMeta(
                        name="id",
                        data_type="integer",
                        udt_name="int4",
                        is_nullable=False,
                        default="",
                        ordinal_position=4,
                        is_identity=True,
                    ),
                },
                conflict_columns=("编号",),
            ),
        }
        example_rows = {
            "美国白蛾调查表": {
                "编号": "MQ0001",
                "受害株数": None,
                "id": 42,
            },
        }

        content = build_import_template_bytes(metadata, example_rows)
        worksheet = load_workbook(BytesIO(content)).active
        self.assertIsNotNone(worksheet)

        examples = [cell.value for cell in worksheet[2]]
        # 真实值覆盖占位
        self.assertEqual(examples[0], "MQ0001")
        # 真实行中的 NULL 留空
        self.assertIsNone(examples[1])
        # 真实行缺少的列回退到类型占位
        self.assertEqual(examples[2], "示例值")
        # 自动生成列即使真实行有值也留空
        self.assertIsNone(examples[3])


class FilterTemplateMetadataTest(unittest.TestCase):
    def _build_metadata(self) -> dict[str, TableMeta]:
        return {
            "美国白蛾调查表": TableMeta(
                schema_name="survey",
                name="美国白蛾调查表",
                columns={},
                conflict_columns=("编号",),
            ),
            "美国白蛾问题点位事件流水表": TableMeta(
                schema_name="ledger",
                name="美国白蛾问题点位事件流水表",
                columns={},
                conflict_columns=("id",),
            ),
            "春尺蠖幼虫调查表": TableMeta(
                schema_name="survey",
                name="春尺蠖幼虫调查表",
                columns={},
                conflict_columns=("编号",),
            ),
        }

    def test_filter_keeps_only_tables_of_the_pest(self) -> None:
        metadata = self._build_metadata()

        filtered = filter_template_metadata(metadata, "美国白蛾")

        self.assertEqual(
            sorted(filtered.keys()),
            ["美国白蛾调查表", "美国白蛾问题点位事件流水表"],
        )

    def test_filter_rejects_unknown_pest_type(self) -> None:
        metadata = self._build_metadata()

        with self.assertRaises(ValueError):
            filter_template_metadata(metadata, "不存在的虫")

    def test_filter_rejects_pest_without_tables(self) -> None:
        with self.assertRaises(ValueError):
            filter_template_metadata({}, "美国白蛾")


if __name__ == "__main__":
    unittest.main()
